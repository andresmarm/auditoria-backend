"""
GENERADOR DE EXCEL — Formato CIC-FT-004
========================================
Llena el formato institucional con el plan generado por Claude.

Mapa de celdas master (celdas combinadas, se escribe en la master):
  A4   → nombre de la auditoría       (rango A4:C4)
  E4   → fecha inicio                 (rango E4:F4)
  H4   → fecha fin                    (celda simple)
  A7   → proceso a auditar            (rango A7:C7)
  D7   → ubicación                    (rango D7:G7)
  H7   → lugar                        (celda simple)
  A9   → antecedentes                 (rango A9:H9)
  A12  → objetivo general             (rango A12:H13)
  A14  → alcance                      (rango A14:H15)
  filas 18-27 → actividades (A=item, B18:D=proceso, E=fecha, F=hora, G=auditado, H=auditor)
  A29  → documentos de referencia     (rango A29:H29)
  A31  → observaciones                (rango A31:H31)
"""

import io, copy, re
from zipfile import ZipFile, ZIP_DEFLATED
from xml.etree import ElementTree as ET
from openpyxl import load_workbook
from openpyxl.styles import Alignment


def llenar_formato_excel(plan_json: dict, template_bytes: bytes) -> bytes:
    """
    Rellena el formato Excel con los datos del plan.
    Args:
        plan_json:      dict con la estructura del plan
        template_bytes: bytes del template .xlsx original
    Returns:
        bytes del Excel relleno
    """
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb.active

    # ── Encabezado ────────────────────────────────────────
    _set(ws, "A4",  plan_json.get("auditoria", ""))
    _set(ws, "E4",  plan_json.get("fecha_inicio", ""))
    _set(ws, "H4",  plan_json.get("fecha_fin", ""))

    # ── Proceso / ubicación / lugar ───────────────────────
    _set(ws, "A7",  plan_json.get("proceso", ""))
    _set(ws, "D7",  plan_json.get("ubicacion", ""))
    _set(ws, "H7",  plan_json.get("lugar", ""))

    # ── Secciones de texto ────────────────────────────────
    _set(ws, "A9",  plan_json.get("antecedentes", ""),      wrap=True)
    _set(ws, "A12", plan_json.get("objetivo_general", ""),  wrap=True)
    _set(ws, "A14", plan_json.get("alcance", ""),           wrap=True)

    # ── Tabla de actividades (filas 18-27) ────────────────
    actividades = plan_json.get("actividades", [])
    for i, act in enumerate(actividades[:10]):   # máx 10 actividades
        fila = 18 + i
        ws.cell(row=fila, column=1).value = i + 1                       # ITEM
        ws.cell(row=fila, column=2).value = act.get("proceso", "")      # PROCESO (master de B:D)
        ws.cell(row=fila, column=5).value = act.get("fecha", "")        # FECHA
        ws.cell(row=fila, column=6).value = act.get("hora", "")         # HORA
        ws.cell(row=fila, column=7).value = act.get("auditado", "")     # AUDITADO
        ws.cell(row=fila, column=8).value = act.get("auditor", "")      # AUDITOR

    # ── Documentos y observaciones ────────────────────────
    _set(ws, "A29", plan_json.get("documentos_referencia", ""), wrap=True)
    _set(ws, "A31", plan_json.get("observaciones", ""),          wrap=True)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return _preservar_recursos_template(template_bytes, out.read())


def _set(ws, coord: str, valor: str, wrap: bool = False):
    """Escribe en la celda master ignorando el error de MergedCell."""
    cell = ws[coord]
    try:
        cell.value = valor
        if wrap:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    except AttributeError:
        # La coord es una celda subordinada del merge; buscar la master
        for rng in ws.merged_cells.ranges:
            from openpyxl.utils import get_column_letter, column_index_from_string
            col_letter = coord[0] if coord[1].isdigit() else coord[:2]
            row = int(coord.lstrip('ABCDEFGHIJKLMNOPQRSTUVWXYZ'))
            col = column_index_from_string(col_letter)
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                from openpyxl.utils import get_column_letter
                master = f"{get_column_letter(rng.min_col)}{rng.min_row}"
                m_cell = ws[master]
                m_cell.value = valor
                if wrap:
                    m_cell.alignment = Alignment(wrap_text=True, vertical="top")
                break


def _preservar_recursos_template(template_bytes: bytes, generated_bytes: bytes) -> bytes:
    """Reinyecta imagenes/drawings que openpyxl puede descartar al guardar."""
    with ZipFile(io.BytesIO(template_bytes), "r") as template_zip:
        template_names = set(template_zip.namelist())
        recursos = [
            name for name in template_names
            if name.startswith(("xl/media/", "xl/drawings/", "xl/printerSettings/"))
        ]
        sheet_rels = "xl/worksheets/_rels/sheet1.xml.rels"
        tiene_drawing = "xl/drawings/drawing1.xml" in template_names and sheet_rels in template_names

        if not recursos and not tiene_drawing:
            return generated_bytes

        salida = io.BytesIO()
        omitidos = set(recursos)
        if tiene_drawing:
            omitidos.add(sheet_rels)

        with ZipFile(io.BytesIO(generated_bytes), "r") as generated_zip:
            with ZipFile(salida, "w", ZIP_DEFLATED) as output_zip:
                for item in generated_zip.infolist():
                    if item.filename in omitidos:
                        continue
                    data = generated_zip.read(item.filename)
                    if item.filename == "xl/worksheets/sheet1.xml" and tiene_drawing:
                        data = _agregar_drawing_a_sheet(data)
                    elif item.filename == "[Content_Types].xml":
                        data = _agregar_content_types(data, template_zip)
                    output_zip.writestr(item, data)

                for name in recursos:
                    output_zip.writestr(name, template_zip.read(name))
                if tiene_drawing:
                    output_zip.writestr(sheet_rels, template_zip.read(sheet_rels))

        return salida.getvalue()


def _agregar_drawing_a_sheet(sheet_xml: bytes) -> bytes:
    texto = sheet_xml.decode("utf-8")
    texto = re.sub(r"<drawing\b[^>]*/>", "", texto)
    if "xmlns:r=" not in texto:
        texto = texto.replace(
            "<worksheet ",
            '<worksheet xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" ',
            1,
        )
    return texto.replace("</worksheet>", '<drawing r:id="rId2"/></worksheet>').encode("utf-8")


def _agregar_content_types(content_xml: bytes, template_zip: ZipFile) -> bytes:
    ns = "http://schemas.openxmlformats.org/package/2006/content-types"
    ET.register_namespace("", ns)
    root = ET.fromstring(content_xml)

    defaults = {el.attrib.get("Extension") for el in root.findall(f"{{{ns}}}Default")}
    overrides = {el.attrib.get("PartName") for el in root.findall(f"{{{ns}}}Override")}

    template_content = ET.fromstring(template_zip.read("[Content_Types].xml"))
    for default in template_content.findall(f"{{{ns}}}Default"):
        extension = default.attrib.get("Extension", "")
        if extension in {"png", "bin"} and extension not in defaults:
            ET.SubElement(root, f"{{{ns}}}Default", default.attrib)

    for override in template_content.findall(f"{{{ns}}}Override"):
        part_name = override.attrib.get("PartName", "")
        if part_name.startswith("/xl/drawings/") and part_name not in overrides:
            ET.SubElement(root, f"{{{ns}}}Override", override.attrib)

    return ET.tostring(root, encoding="utf-8", xml_declaration=False)


# ── PROMPT JSON para Claude ───────────────────────────────

PROMPT_JSON_PLAN = """
Genera un plan de auditoría en formato JSON ESTRICTO.
Responde ÚNICAMENTE con el JSON, sin explicaciones ni bloques de código markdown.

Estructura exacta requerida:
{
  "auditoria": "Nombre descriptivo de la auditoría",
  "fecha_inicio": "YYYY-MM-DD",
  "fecha_fin": "YYYY-MM-DD",
  "proceso": "Nombre del proceso a auditar",
  "ubicacion": "Dependencia o área responsable",
  "lugar": "Ubicación física (ej: Sede Principal COMCAJA)",
  "antecedentes": "Contexto y antecedentes en 2-3 párrafos",
  "objetivo_general": "Objetivo general de la auditoría",
  "alcance": "Alcance: procesos, períodos, dependencias involucradas",
  "actividades": [
    {
      "proceso": "Descripción concreta de la actividad (máx 120 caracteres)",
      "fecha": "YYYY-MM-DD",
      "hora": "HH:MM",
      "auditado": "Cargo o área auditada",
      "auditor": "Nombre o cargo del auditor"
    }
  ],
  "documentos_referencia": "Normas citadas: LEY XX ART. Y, DECRETO ZZ...",
  "observaciones": "Metodología y observaciones adicionales"
}

Genera SOLO las actividades de auditoría necesarias y suficientes para cubrir el procedimiento real.
NO rellenes con actividades genéricas para llegar a un número fijo.
- Un procedimiento sencillo: 4-6 actividades.
- Un procedimiento complejo: 7-9 actividades.
- Dos o más procedimientos combinados: distribuye las actividades entre ellos, indica en "proceso" a cuál pertenece cada una. Máximo 10 en total.
- La ÚLTIMA actividad de la lista debe ser siempre la de comunicación de resultados
  (mesa de cierre, presentación de hallazgos y acciones de mejora con el auditado).
Usa fechas realistas comenzando 15 días desde hoy si no se especifica un período de ejecución.
"""


# ── Test local ────────────────────────────────────────────

if __name__ == "__main__":
    plan = {
        "auditoria": "Auditoría Proceso Subsidio Familiar de Vivienda",
        "fecha_inicio": "2025-10-01",
        "fecha_fin": "2025-10-15",
        "proceso": "Gestión del Subsidio Familiar de Vivienda",
        "ubicacion": "Oficina de Subsidios",
        "lugar": "Sede Principal COMCAJA",
        "antecedentes": "Auditoría programada en el Plan Anual 2025 según Ley 87 de 1993.",
        "objetivo_general": "Evaluar el cumplimiento del Decreto 1077 de 2015 en la asignación del SVF.",
        "alcance": "Solicitudes aprobadas período enero-junio 2025. Áreas: Subsidios, Jurídica, Financiera.",
        "actividades": [
            {"proceso": "Revisión expedientes postulantes - verificación requisitos Decreto 1077/2015", "fecha": "2025-10-01", "hora": "08:00", "auditado": "Coordinador de Subsidios", "auditor": "Auditor Interno"},
            {"proceso": "Verificación elegibilidad beneficiarios Art. 3 Ley 21/1982", "fecha": "2025-10-02", "hora": "08:00", "auditado": "Analista de Subsidios", "auditor": "Auditor Interno"},
            {"proceso": "Revisión actos administrativos y soportes jurídicos de asignación", "fecha": "2025-10-03", "hora": "09:00", "auditado": "Jefe Oficina Jurídica", "auditor": "Auditor Interno"},
            {"proceso": "Verificación pagos y conciliación registros financieros", "fecha": "2025-10-06", "hora": "08:00", "auditado": "Tesorero", "auditor": "Auditor Interno"},
            {"proceso": "Revisión reportes SSF - Circular Externa 2016-00020", "fecha": "2025-10-07", "hora": "09:00", "auditado": "Jefe de Subsidios", "auditor": "Auditor Interno"},
            {"proceso": "Mesa de cierre - presentación hallazgos preliminares", "fecha": "2025-10-14", "hora": "14:00", "auditado": "Gerencia y Jefes de área", "auditor": "Jefe Control Interno"},
        ],
        "documentos_referencia": "LEY 21 DE 1982 | DECRETO 1077 DE 2015 | RESOLUCION 610 DE 2004 | LEY 87 DE 1993",
        "observaciones": "Metodología: visita de rutina. Técnicas: inspección documental, entrevistas y muestreo."
    }

    template = open('/mnt/user-data/uploads/CIC-FT-004_FORMATO_PLAN_DE_TRABAJO_DE_AUDITORIA.xlsx','rb').read()
    resultado = llenar_formato_excel(plan, template)
    with open('/home/claude/plan_test.xlsx','wb') as f:
        f.write(resultado)
    print(f"✓ Excel generado: {len(resultado):,} bytes → /home/claude/plan_test.xlsx")
